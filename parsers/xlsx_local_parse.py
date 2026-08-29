"""XLSX 로컬 결정론 파서 (docparse Tier 0): openpyxl로 외부 API 없이 추출.

다른 *_parse.py와 동일한 인터페이스:
  python xlsx_local_parse.py "<파일.xlsx>"        → <파일>_xlsxlocal.md

XLSX는 Tier 0 철학의 최적 대상이다: 셀 값·병합 범위·숨김 상태가 파일 안에
명시돼 있어 파싱이 추론이 아니라 읽기다(괘선 PDF보다도 명시적). 병합 셀은
범위가 정확히 기록돼 있으므로 거부하지 않고 앵커 셀에 값 + 범위를 명시
기록한다(span 손실 없음). 숨김 행·열의 데이터도 보존하고 고지한다(시각적
숨김이 데이터 부재가 아니다). 미계산 수식(캐시값 없음)은 수식 원문을 그대로
보존하고 고지한다(계산하지 않는다 — 추론 금지).

독립 교차 검증 내장: openpyxl 추출값 전체를 원시 XML 직접 파스(zipfile +
ElementTree로 sheet XML·sharedStrings를 자체 해석)와 값 멀티셋으로 대조한다
(crosscheck_with_rawxml). 두 독립 구현이 같은 값 집합에 도달해야만 PASS.
불일치가 하나라도 있으면 출력 파일을 만들지 않는다 — 기존 티어(Upstage·
LlamaParse)로 승격할 것. 날짜 셀은 openpyxl 변환값을 시리얼로 되돌려
대조하므로 이 부분만 라이브러리 유틸을 공유한다(한계 명시). 미계산 수식의
수식 원문 표기는 값이 아니므로 이 대조의 대상이 아니다(한계 명시: 공유 수식
si 상속 해석이 필요해 검증하지 않는다).

한계: 차트·이미지 안의 텍스트는 범위 밖(존재 시 경고, 출력은 유지 — hwpx_local
과 동일 정책). xlsm(매크로)·암호화 파일 미지원. 셀 서식(색·테두리)은 데이터가
아니므로 다루지 않는다.

셀 밖 텍스트(2026-08-31): 두 검증 경로가 모두 셀 값만 보므로, 별도 XML에 저장되는
텍스트는 대조로 잡히지 않는다. ① 도형·텍스트박스(xl/drawings/*.xml의 a:t)는 시트에
보이는 본문이므로 존재 시 **거부**한다(docx_local의 텍스트박스 거부와 같은 정책).
② 셀 주석·메모(xl/comments*.xml, xl/threadedComments/)는 본문이 아니므로 출력에
넣지 않되 **개수를 고지**한다(docx_local의 검토 메모 고지와 같은 정책).

종료 코드: 0 = PASS(출력 파일 생성), 1 = 오류·거부(출력 없음). 실행 시작 시
같은 이름의 이전 출력을 지운다.
"""
import os
import sys
import traceback
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter
from datetime import date, datetime, time, timedelta

# Windows cp949 안전 (다른 파서와 동일 패턴: 양쪽 스트림 + replace)
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

SUPPORTED_EXTENSIONS = [".xlsx"]
MAX_PROBLEM_SAMPLES = 10
NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
NS_DOCREL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
NS_PKGREL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
NS_A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"


def get_output_filename(input_file):
    """입력 파일 경로 기반 출력 경로(<base>_xlsxlocal.md) 생성."""
    dir_path = os.path.dirname(input_file)
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    output_name = f"{base_name}_xlsxlocal.md"
    return os.path.join(dir_path, output_name) if dir_path else output_name


def clear_stale_output(output_file):
    """같은 이름의 이전 출력이 있으면 지운다(실패한 실행 뒤 옛 파일 잔존 방지)."""
    if os.path.exists(output_file):
        os.remove(output_file)
        print(f"기존 출력 제거: {output_file}")


def scan_offcell_text(zf, names):
    """셀 밖 텍스트 탐지: (도형 텍스트 샘플 목록, 주석 개수).

    도형·텍스트박스 텍스트는 xl/drawings/drawingN.xml의 a:t에 있다(차트 파트
    xl/charts/는 별도 경고 대상이라 제외). 주석은 xl/comments*.xml의 t 요소와
    xl/threadedComments/*.xml의 threadedComment 요소를 센다.
    """
    shape_texts = []
    for n in names:
        if not (n.startswith("xl/drawings/") and n.endswith(".xml")):
            continue
        try:
            root = ET.fromstring(zf.read(n))
        except ET.ParseError:
            continue
        for t in root.iter(f"{NS_A}t"):
            if t.text and t.text.strip():
                shape_texts.append(t.text.strip())
    comments = 0
    for n in names:
        base = os.path.basename(n).lower()
        # Excel은 xl/comments1.xml, openpyxl은 xl/comments/comment1.xml로 쓴다. 둘 다 잡는다.
        if n.startswith("xl/") and "comment" in base and n.endswith(".xml") and "threaded" not in n:
            try:
                root = ET.fromstring(zf.read(n))
            except ET.ParseError:
                continue
            comments += sum(1 for _ in root.iter(f"{NS_MAIN}comment"))
        elif n.startswith("xl/threadedComments/") and n.endswith(".xml"):
            try:
                root = ET.fromstring(zf.read(n))
            except ET.ParseError:
                continue
            comments += sum(1 for el in root.iter() if el.tag.endswith("}threadedComment"))
    return shape_texts, comments


def _comparable(value, epoch):
    """openpyxl 셀 값을 원시 XML 값과 대조 가능한 표준형으로. None이면 값 없음."""
    from openpyxl.utils.datetime import to_excel, time_to_days, timedelta_to_days

    if value is None:
        return None
    if isinstance(value, bool):
        return ("bool", value)
    if isinstance(value, str):
        return ("text", value)
    if isinstance(value, (int, float)):
        return ("num", round(float(value), 7))
    if isinstance(value, (datetime, date)):
        return ("num", round(float(to_excel(value, epoch)), 7))
    if isinstance(value, time):
        return ("num", round(float(time_to_days(value)), 7))
    if isinstance(value, timedelta):
        return ("num", round(float(timedelta_to_days(value)), 7))
    return ("text", str(value))


def collect_pyxl_values(filename):
    """openpyxl(data_only) 관점의 시트별 값 목록 [(시트명, [표준형, ...]), ...]."""
    import openpyxl

    wb = openpyxl.load_workbook(filename, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        values = []
        for row in ws.iter_rows():
            for cell in row:
                cv = _comparable(cell.value, wb.epoch)
                if cv is not None:
                    values.append(cv)
        sheets.append((ws.title, values))
    wb.close()
    return sheets


def _shared_strings(zf):
    """sharedStrings.xml → 인덱스별 텍스트(리치 텍스트 run 연결, 발음 주석 제외)."""
    strings = []
    try:
        data = zf.read("xl/sharedStrings.xml")
    except KeyError:
        return strings
    root = ET.fromstring(data)
    for si in root.findall(f"{NS_MAIN}si"):
        parts = []
        for t in si.iter(f"{NS_MAIN}t"):
            # rPh(발음 주석) 안의 t는 openpyxl도 제외하므로 대조에서 뺀다
            skip = False
            for rph in si.iter(f"{NS_MAIN}rPh"):
                if t in rph.iter(f"{NS_MAIN}t"):
                    skip = True
                    break
            if not skip:
                parts.append(t.text or "")
        strings.append("".join(parts))
    return strings


def _sheet_files(zf):
    """workbook.xml + rels로 시트 순서대로 (시트명, XML 경로) 매핑."""
    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {}
    for rel in rels_root.findall(f"{NS_PKGREL}Relationship"):
        rid_to_target[rel.get("Id")] = rel.get("Target")
    result = []
    for sheet in wb_root.iter(f"{NS_MAIN}sheet"):
        name = sheet.get("name")
        rid = sheet.get(f"{NS_DOCREL}id")
        target = rid_to_target.get(rid, "")
        if target.startswith("/"):
            path = target.lstrip("/")
        else:
            path = "xl/" + target
        result.append((name, path.replace("\\", "/")))
    return result


def _raw_sheet_values(zf, sheet_path, strings):
    """시트 XML을 자체 해석해 값 표준형 목록으로 (openpyxl과 독립적인 관점)."""
    values = []
    root = ET.fromstring(zf.read(sheet_path))
    for c in root.iter(f"{NS_MAIN}c"):
        ctype = c.get("t", "n")
        v = c.find(f"{NS_MAIN}v")
        if ctype == "inlineStr":
            is_el = c.find(f"{NS_MAIN}is")
            if is_el is not None:
                text = "".join(t.text or "" for t in is_el.iter(f"{NS_MAIN}t"))
                values.append(("text", text))
            continue
        if v is None or v.text is None:
            continue  # 값 없는 셀(서식만) 또는 미계산 수식
        raw = v.text
        if ctype == "s":
            values.append(("text", strings[int(raw)]))
        elif ctype == "str":
            values.append(("text", raw))
        elif ctype == "b":
            values.append(("bool", raw in ("1", "true", "TRUE")))
        elif ctype == "e":
            values.append(("text", raw))  # 오류값(#DIV/0! 등)은 문자로 대조
        elif ctype == "d":
            # ISO 8601 날짜 셀(희귀 작성기): 시리얼로 환산해 openpyxl 관점과 대조
            try:
                from datetime import datetime as _dt
                from openpyxl.utils.datetime import to_excel
                values.append(("num", round(float(to_excel(_dt.fromisoformat(raw))), 7)))
            except Exception:
                values.append(("text", raw))
        else:
            values.append(("num", round(float(raw), 7)))
    return values


def crosscheck_with_rawxml(filename, sheets_values):
    """openpyxl 추출값을 원시 XML 자체 해석과 시트 단위 멀티셋으로 교차 검증.

    sheets_values: collect_pyxl_values() 형식 [(시트명, [표준형, ...]), ...]
    반환: (대조 값 수, 문제 목록[str])
    """
    checked = 0
    problems = []
    with zipfile.ZipFile(filename) as zf:
        strings = _shared_strings(zf)
        raw_sheets = _sheet_files(zf)
        raw_names = [n for n, _ in raw_sheets]
        pyxl_names = [n for n, _ in sheets_values]
        if raw_names != pyxl_names:
            problems.append(f"시트 목록 불일치: openpyxl {pyxl_names} / XML {raw_names}")
            return checked, problems
        for (name, path), (_, pyxl_vals) in zip(raw_sheets, sheets_values):
            raw_vals = _raw_sheet_values(zf, path, strings)
            checked += len(raw_vals)
            diff_raw = Counter(raw_vals) - Counter(pyxl_vals)
            diff_pyxl = Counter(pyxl_vals) - Counter(raw_vals)
            for val, cnt in list(diff_raw.items())[:MAX_PROBLEM_SAMPLES]:
                problems.append(f"시트 '{name}': XML에만 있는 값 {val!r} x{cnt}")
            for val, cnt in list(diff_pyxl.items())[:MAX_PROBLEM_SAMPLES]:
                problems.append(f"시트 '{name}': openpyxl에만 있는 값 {val!r} x{cnt}")
    return checked, problems


def _fmt_cell(value):
    """셀 값을 마크다운 표기용 문자열로 (결정론적 표기, 데이터 불변)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.hour == value.minute == value.second == value.microsecond == 0:
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return " ".join(str(value).split()).replace("|", "\\|")


def _grid_to_markdown(grid):
    if not grid:
        return ""
    width = max(len(row) for row in grid)
    lines = []
    for i, row in enumerate(grid):
        cells = [_fmt_cell(v) for v in row] + [""] * (width - len(row))
        lines.append("| " + " | ".join(cells) + " |")
        if i == 0:
            lines.append("|" + " --- |" * width)
    return "\n".join(lines)


def _used_bounds(ws):
    """실제 값이 있는 마지막 행·열 (max_row/max_column의 빈 서식 영역 과대 방지)."""
    max_r = 0
    max_c = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_r = max(max_r, cell.row)
                max_c = max(max_c, cell.column)
    return max_r, max_c


def process_file(filename):
    """PASS로 출력 파일을 만들었으면 True."""
    import openpyxl
    from openpyxl.utils import column_index_from_string, get_column_letter

    print(f"입력 파일: {filename}")
    ext = os.path.splitext(filename)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        print(f"오류: 지원하지 않는 파일 형식입니다: {ext}")
        if ext == ".xls":
            print("  구형 XLS는 Excel/LibreOffice로 XLSX 저장 후 입력하세요.")
        elif ext == ".xlsm":
            print("  매크로 포함 XLSM은 XLSX로 저장 후 입력하세요.")
        print(f"지원 형식: {', '.join(SUPPORTED_EXTENSIONS)}")
        return False
    if not os.path.exists(filename):
        print(f"오류: 파일을 찾을 수 없습니다: {filename}")
        return False

    output_file = get_output_filename(filename)
    clear_stale_output(output_file)

    print("추출 중... (openpyxl 명시 데이터 + 원시 XML 독립 대조)")
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        wb_formula = openpyxl.load_workbook(filename, data_only=False)
    except Exception as e:
        print(f"경고: openpyxl이 통합 문서를 읽지 못했습니다: {e}")
        print("→ 차트시트·특수 구조 가능성. 출력 파일을 만들지 않았습니다. 기존 티어(Upstage·LlamaParse)로 승격하세요.")
        return False

    sections = []
    total_cells = 0
    uncached_formulas = 0
    cached_formulas = 0
    merged_total = 0
    hidden_notes = []

    for ws, ws_f in zip(wb.worksheets, wb_formula.worksheets):
        # 경계는 수식 관점(ws_f)으로 계산: 미계산 수식 셀은 data_only 관점에서
        # 값이 None이라 경계 계산에서 행·열째 잘려 나간다.
        max_r, max_c = _used_bounds(ws_f)
        merged_anchors = {}
        merged_members = set()
        for rng in ws.merged_cells.ranges:
            merged_anchors[(rng.min_row, rng.min_col)] = str(rng)
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    merged_members.add((r, c))
            max_r = max(max_r, rng.max_row)
            max_c = max(max_c, rng.max_col)
        merged_total += len(ws.merged_cells.ranges)

        grid = []
        for r in range(1, max_r + 1):
            row_vals = []
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                value = cell.value
                if value is not None:
                    total_cells += 1
                cell_f = ws_f.cell(row=r, column=c)
                if cell_f.data_type == "f":
                    if value is None:
                        # 미계산 수식: 계산하지 않고 수식 원문을 보존한다
                        value = cell_f.value
                        uncached_formulas += 1
                    else:
                        cached_formulas += 1
                row_vals.append(value)
            grid.append(row_vals)

        body = []
        if len(wb.worksheets) > 1:
            state_note = " (숨김 시트)" if ws.sheet_state != "visible" else ""
            body.append(f"## 시트: {ws.title}{state_note}")
        md = _grid_to_markdown(grid)
        if md:
            body.append(md)
        else:
            body.append("(빈 시트)")
        if merged_anchors:
            body.append("병합 셀: " + ", ".join(sorted(merged_anchors.values())))
        hidden_rows = sorted(
            r for r, dim in ws.row_dimensions.items() if dim.hidden and r <= max_r
        )
        # 열 dim은 그룹 숨김 시 min~max 범위 하나로 저장되므로 전체 범위를 전개한다
        hidden_col_idx = set()
        for dim in ws.column_dimensions.values():
            if not dim.hidden:
                continue
            lo = dim.min or column_index_from_string(dim.index)
            hi = dim.max or lo
            for idx in range(lo, hi + 1):
                if idx <= max_c:
                    hidden_col_idx.add(idx)
        hidden_cols = [get_column_letter(i) for i in sorted(hidden_col_idx)]
        if hidden_rows or hidden_cols:
            parts = []
            if hidden_rows:
                parts.append("행 " + ", ".join(str(r) for r in hidden_rows))
            if hidden_cols:
                parts.append("열 " + ", ".join(hidden_cols))
            note = "숨김(데이터는 보존됨): " + " · ".join(parts)
            body.append(note)
            hidden_notes.append(f"{ws.title}: {note}")
        sections.append("\n\n".join(body))

    sheet_count = len(wb.worksheets)
    wb.close()
    wb_formula.close()

    with zipfile.ZipFile(filename) as zf:
        names = zf.namelist()
        shape_texts, comment_count = scan_offcell_text(zf, names)
    media = sum(1 for n in names if n.startswith("xl/media/"))
    charts = sum(1 for n in names if n.startswith("xl/charts/chart"))

    print(f"시트 {sheet_count}개 · 값 있는 셀 {total_cells}개 · 병합 범위 {merged_total}개")
    if uncached_formulas:
        print(f"[알림] 미계산 수식 {uncached_formulas}건: 캐시값이 없어 수식 원문을 보존했습니다(계산하지 않음).")
    if cached_formulas:
        print(f"[알림] 수식 셀 {cached_formulas}건은 저장된 캐시값을 사용했습니다.")
    for note in hidden_notes:
        print(f"[알림] {note}")

    sheets_values = collect_pyxl_values(filename)
    checked, problems = crosscheck_with_rawxml(filename, sheets_values)
    print(f"교차 검증(openpyxl / 원시 XML 독립 해석): 값 {checked}건 대조")

    if shape_texts:
        sample = " / ".join(t[:30] for t in shape_texts[:5])
        problems.append(
            f"도형·텍스트박스 텍스트 {len(shape_texts)}건(셀 밖, 두 검증 경로 모두 범위 밖): {sample}"
        )

    if problems:
        print(f"경고: 검증 문제 {len(problems)}건")
        for line in problems[:MAX_PROBLEM_SAMPLES]:
            print(f"  - {line}")
        if len(problems) > MAX_PROBLEM_SAMPLES:
            print(f"  ... 외 {len(problems) - MAX_PROBLEM_SAMPLES}건")
        print("→ 검증 실패: 출력 파일을 만들지 않았습니다. 기존 티어(Upstage·LlamaParse)로 승격하세요.")
        return False

    if comment_count:
        print(f"[알림] 셀 주석·메모 {comment_count}건이 있습니다. 셀 값이 아니므로 출력에 포함하지 않았습니다.")
    if media or charts:
        print(f"경고: 이미지 {media}개 · 차트 {charts}개 존재. 그 안의 텍스트는 추출 범위 밖입니다.")
        print("  이미지 내 텍스트가 중요하면 Upstage(upstage_parse.py)로 교차검증하세요.")

    with open(output_file, "w", encoding="utf-8") as f:
        f.write("\n\n".join(s for s in sections if s) + "\n")
    print(f"출력 파일: {output_file}")
    print("=== PASS: 원시 XML 독립 대조까지 완전 일치 ===")
    return True


def main():
    """반환값이 종료 코드다(0 성공 / 1 실패)."""
    try:
        import openpyxl  # noqa: F401
    except ImportError as e:
        print("오류: openpyxl 패키지를 찾을 수 없습니다.")
        print("설치 명령: pip install openpyxl")
        print(f"상세: {e}")
        return 1

    positional = [a for a in sys.argv[1:] if not a.startswith("--")]
    if positional:
        return 0 if process_file(positional[0]) else 1

    supported_files = sorted(
        f for f in os.listdir(".")
        if os.path.isfile(f) and os.path.splitext(f)[1].lower() in SUPPORTED_EXTENSIONS
    )
    if not supported_files:
        print("오류: 현재 디렉토리에 지원되는 XLSX 파일이 없습니다.")
        print(f"지원 형식: {', '.join(SUPPORTED_EXTENSIONS)}")
        return 1
    if len(supported_files) == 1:
        return 0 if process_file(supported_files[0]) else 1
    print(f"XLSX 파일 {len(supported_files)}개 발견:")
    for i, f in enumerate(supported_files, 1):
        print(f"  {i}. {f}")
    print()
    all_ok = True
    for f in supported_files:
        all_ok = process_file(f) and all_ok
        print()
    return 0 if all_ok else 1


if __name__ == "__main__":
    exit_code = 1
    try:
        exit_code = main()
    except Exception as e:
        print(f"\n오류 발생: {e}")
        print("\n상세 정보:")
        traceback.print_exc()

    # 비-TTY(AI 에이전트·파이프·백그라운드)에서는 stdin이 EOF로 닫히지 않아
    # input()이 무한 블록되므로 isatty()로 가드 (다른 *_parse.py와 동일 패턴).
    if sys.stdin is not None and sys.stdin.isatty():
        try:
            input("\nEnter를 눌러 종료...")
        except EOFError:
            pass
    sys.exit(exit_code)
